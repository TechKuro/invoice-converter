#!/usr/bin/env python3
"""
PDF to Excel Converter
======================

A Python script that extracts data from multiple PDF files and exports the results to an Excel spreadsheet.
Supports multiple extraction methods including text extraction, table detection, and structured data parsing.

Usage:
    python pdf_to_excel_converter.py [options]

Examples:
    python pdf_to_excel_converter.py --input-dir ./pdfs --output-file results.xlsx
    python pdf_to_excel_converter.py --input-dir ./invoices --extract-tables --output-file invoice_data.xlsx
"""

import os
import sys
import argparse
import logging
import string
from pathlib import Path
from typing import List, Dict, Any, Optional
import re
from datetime import datetime
import subprocess

# Add this block right after standard library imports and before third-party imports
def _ensure_dependencies():
    """Install any missing packages listed in requirements.txt at runtime."""
    req_file = Path(__file__).with_name('requirements.txt')
    if not req_file.exists():
        return  # nothing to do
    try:
        import pkg_resources  # part of setuptools
    except ImportError:
        # setuptools is almost always present with pip; if not, try to install
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'setuptools'])
        import pkg_resources
    
    installed = {pkg.key for pkg in pkg_resources.working_set}
    missing = []
    with req_file.open() as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            pkg_name = line.split('==')[0].lower()
            if pkg_name not in installed:
                missing.append(line)
    if missing:
        print('Installing missing dependencies:', ', '.join(missing))
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', *missing])

# Ensure dependencies are present before imports below
_ensure_dependencies()

# Third-party imports
try:
    import pandas as pd
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please install required packages: pip install -r requirements.txt")
    sys.exit(1)


class PDFDataExtractor:
    """Handles extraction of data from PDF files."""
    
    def __init__(self, extract_tables: bool = True, extract_text: bool = True):
        self.extract_tables = extract_tables
        self.extract_text = extract_text
        self.logger = logging.getLogger(__name__)
    
    def extract_from_pdf(self, pdf_path: Path) -> Dict[str, Any]:
        """
        Extract data from a single PDF file.
        
        Args:
            pdf_path: Path to the PDF file
            
        Returns:
            Dictionary containing extracted data
        """
        extracted_data = {
            'filename': pdf_path.name,
            'filepath': str(pdf_path),
            'text': '',
            'tables': [],
            'metadata': {},
            'extraction_date': datetime.now().isoformat()
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Extract metadata
                if pdf.metadata:
                    extracted_data['metadata'] = {
                        'title': pdf.metadata.get('Title', ''),
                        'author': pdf.metadata.get('Author', ''),
                        'creator': pdf.metadata.get('Creator', ''),
                        'creation_date': str(pdf.metadata.get('CreationDate', '')),
                        'pages': len(pdf.pages)
                    }
                
                all_text = []
                all_tables = []
                
                for page_num, page in enumerate(pdf.pages, 1):
                    if self.extract_text:
                        page_text = page.extract_text()
                        if page_text:
                            all_text.append(f"Page {page_num}:\n{page_text}\n")
                    
                    if self.extract_tables:
                        # Extract tables
                        tables = []
                        
                        # First try default table extraction
                        default_tables = page.extract_tables()
                        if default_tables:
                            for i, table in enumerate(default_tables):
                                if table:  # Skip empty tables
                                    tables.append(table)
                                    all_tables.append({
                                        'page': page_num,
                                        'table_number': i + 1,
                                        'data': table,
                                        'rows': len(table),
                                        'columns': len(table[0]) if table else 0
                                    })
                        
                        # If no tables found with default method, try text-based detection
                        if not tables:
                            table_settings = {
                                "vertical_strategy": "text",
                                "horizontal_strategy": "text", 
                                "intersection_tolerance": 5,
                                "min_words_vertical": 1,
                                "min_words_horizontal": 1
                            }
                            text_tables = page.extract_tables(table_settings)
                            
                            for i, table in enumerate(text_tables):
                                if table and len(table) > 1:  # Skip empty or single-row tables
                                    # Filter out mostly empty tables
                                    meaningful_rows = 0
                                    for row in table:
                                        if row and any(cell and str(cell).strip() for cell in row):
                                            meaningful_rows += 1
                                    
                                    if meaningful_rows >= 2:  # At least 2 meaningful rows
                                        tables.append(table)
                                        all_tables.append({
                                            'page': page_num,
                                            'table_number': i + 1,
                                            'data': table,
                                            'rows': len(table),
                                            'columns': len(table[0]) if table else 0,
                                            'detection_method': 'text-based'
                                        })
                
                extracted_data['text'] = '\n'.join(all_text)
                extracted_data['tables'] = all_tables
                
                # Extract line items from tables and text
                extracted_data['line_items'] = self._extract_line_items_from_data(all_tables, extracted_data['text'])
                
        except Exception as e:
            self.logger.error(f"Error extracting data from {pdf_path}: {str(e)}")
            extracted_data['error'] = str(e)
        
        return extracted_data
    
    def extract_invoice_data(self, text: str) -> Dict[str, str]:
        """
        Extract common invoice fields from text using regex patterns.
        
        Args:
            text: Extracted text from PDF
            
        Returns:
            Dictionary with extracted invoice fields
        """
        invoice_data = {}
        
        # Common patterns for invoice data
        patterns = {
            'invoice_number': [
                r'invoice\s*#?\s*:?\s*([A-Z0-9\-_]+)',
                r'inv\s*#?\s*:?\s*([A-Z0-9\-_]+)',
                r'invoice\s*number\s*:?\s*([A-Z0-9\-_]+)'
            ],
            'date': [
                r'date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'invoice\s*date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'
            ],
            'total_amount': [
                r'total\s*:?\s*\$?([0-9,]+\.?\d*)',
                r'amount\s*due\s*:?\s*\$?([0-9,]+\.?\d*)',
                r'grand\s*total\s*:?\s*\$?([0-9,]+\.?\d*)'
            ],
            'vendor': [
                r'from\s*:?\s*([^\n]+)',
                r'vendor\s*:?\s*([^\n]+)',
                r'company\s*:?\s*([^\n]+)'
            ]
        }
        
        for field, pattern_list in patterns.items():
            for pattern in pattern_list:
                match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if match:
                    invoice_data[field] = match.group(1).strip()
                    break
        
        return invoice_data
    

    
    def _parse_table_for_line_items(self, table_data: List[List], table_info: Dict) -> List[Dict[str, Any]]:
        """Parse a table to extract line items, preserving original column structure."""
        line_items = []
        
        if not table_data or len(table_data) < 2:
            return line_items
        
        # Find header row by looking for table column patterns  
        header_row_idx = -1
        column_headers = []
        
        # Look for the row that seems most like headers
        for row_idx, row in enumerate(table_data[:30]):  # Check first 30 rows for headers
            if not row:
                continue
            
            # Clean and check the row
            cleaned_row = [str(cell).strip() if cell else '' for cell in row]
            non_empty_cells = [cell for cell in cleaned_row if cell and len(cell) > 0]
            
            if len(non_empty_cells) >= 3:  # At least 3 columns with data for table headers
                # Join cells to handle split headers like "Description Quantit" + "y" 
                combined_text = ' '.join(non_empty_cells).lower()
                
                # Must have multiple table column indicators (not just single words)
                table_column_indicators = [
                    'description', 'quantity', 'price', 'amount', 'vat', 'total'
                ]
                column_score = sum(1 for indicator in table_column_indicators if indicator in combined_text)
                
                # Special patterns for invoice tables (handles split headers)
                invoice_patterns = [
                    'quantit',      # "Description Quantit" + "y"
                    'unitprice',    # "UnitPrice"  
                    'amountgbp',    # "AmountGBP"
                    'amount gbp'    # "Amount GBP"
                ]
                invoice_score = sum(1 for pattern in invoice_patterns if pattern in combined_text)
                
                # Check for typical table structure (avoid document headers)
                has_table_structure = (
                    ('description' in combined_text or 'quantit' in combined_text) and
                    ('price' in combined_text or 'amount' in combined_text) and
                    len(non_empty_cells) >= 4  # At least 4 columns
                )
                
                # Detect headers only if we have strong table indicators
                if (column_score >= 2 or invoice_score >= 2 or has_table_structure):
                    header_row_idx = row_idx
                    column_headers = cleaned_row
                    # Debug output to help troubleshoot
                    self.logger.debug(f"Table headers detected in row {row_idx}: {column_headers}")
                    self.logger.debug(f"Combined text: {combined_text}")
                    self.logger.debug(f"Column score: {column_score}, Invoice score: {invoice_score}, Has structure: {has_table_structure}")
                    break
        
        # If no clear header found, create generic headers
        if header_row_idx == -1 and table_data:
            header_row_idx = 0
            first_row = table_data[0] if table_data else []
            column_headers = [f"Column_{i+1}" for i in range(len(first_row))]
            self.logger.debug(f"No headers detected, using generic headers: {column_headers}")
        
        # Process data rows if we have headers
        if header_row_idx >= 0 and column_headers:
            self.logger.debug(f"Processing {len(table_data) - header_row_idx - 1} data rows with headers: {column_headers}")
            for row_idx in range(header_row_idx + 1, len(table_data)):
                row = table_data[row_idx]
                if not row:
                    continue
                
                # Create line item with all columns preserved
                line_item = {
                    'page': table_info['page'],
                    'table_number': table_info['table_number'],
                    'row_number': row_idx - header_row_idx
                }
                
                # Extract data for each column
                has_meaningful_data = False
                for col_idx, header in enumerate(column_headers):
                    if col_idx < len(row) and row[col_idx]:
                        cell_value = str(row[col_idx]).strip()
                        
                        if cell_value and cell_value != '':
                            # Clean header name for use as dictionary key
                            clean_header = self._clean_header_name(header)
                            if not clean_header:
                                clean_header = f"column_{col_idx + 1}"
                            
                            line_item[clean_header] = cell_value
                            
                            # Check if this cell contains meaningful data (not just formatting)
                            if len(cell_value) > 1 and not cell_value.isspace():
                                has_meaningful_data = True
                
                # Skip rows that look like totals, subtotals, or separators
                if has_meaningful_data and not self._is_total_row(line_item):
                    # Clean up the line item before adding
                    cleaned_item = self._clean_line_item(line_item, column_headers)
                    if cleaned_item:  # Only add if cleaning was successful
                        line_items.append(cleaned_item)
        
        return line_items
    
    def _clean_header_name(self, header: str) -> str:
        """Clean header name to be used as a dictionary key."""
        if not header:
            return ''
        
        # Remove special characters and spaces, convert to lowercase
        cleaned = ''.join(c.lower() if c.isalnum() else '_' for c in header)
        # Remove multiple underscores and leading/trailing underscores
        cleaned = re.sub(r'_+', '_', cleaned).strip('_')
        
        # Truncate if too long
        if len(cleaned) > 30:
            cleaned = cleaned[:30]
        
        return cleaned if cleaned else 'column'
    
    def _is_likely_data(self, cell_value: str) -> bool:
        """Check if a cell value looks like data rather than a header."""
        if not cell_value:
            return False
        
        # Numbers or monetary values are likely data
        if re.search(r'[\d\$£€¥%]', cell_value):
            return True
        
        # Very long strings are likely data
        if len(cell_value) > 30:
            return True
        
        # Dates are likely data
        if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', cell_value):
            return True
        
        # Common data patterns that shouldn't be headers
        data_patterns = [
            r'^\d+\.\d+$',  # Decimal numbers
            r'^\d+$',       # Whole numbers
            r'hosted',      # Service descriptions
            r'email',
            r'fax'
        ]
        
        cell_lower = cell_value.lower()
        for pattern in data_patterns:
            if re.search(pattern, cell_lower):
                return True
        
        return False
    
    def _is_total_row(self, row_data: Dict[str, Any]) -> bool:
        """Check if a row represents a total/summary rather than a line item."""
        
        # Convert all values to strings for checking
        text_values = []
        for key, value in row_data.items():
            if key not in ['page', 'table_number', 'row_number'] and value:
                text_values.append(str(value).strip().upper())
        
        combined_text = ' '.join(text_values)
        
        # Check for total/summary indicators
        total_indicators = [
            'TOTAL', 'SUBTOTAL', 'SUB-TOTAL', 'GRAND TOTAL',
            'VAT', 'TAX', 'NET', 'GROSS', 'BALANCE',
            'AMOUNT DUE', 'INVOICE TOTAL', 'FINAL TOTAL',
            'SHIPPING', 'DELIVERY', 'DISCOUNT',
            'TOTA', 'OTAL',  # Partial matches for split text
        ]
        
        # Check if any value contains total indicators
        for indicator in total_indicators:
            if indicator in combined_text:
                return True
        
        # Check for rows with mostly empty fields or single words
        meaningful_fields = [v for v in text_values if len(v) > 2]
        if len(meaningful_fields) <= 1:
            return True
        
        # Check for company/header information
        company_indicators = [
            'LIMITED', 'LTD', 'LLC', 'CORP', 'CORPORATION', 'INC',
            'TECHNOLOGY', 'SOLUTIONS', 'SERVICES', 'GROUP',
            'COMPANY', 'ANEXIAN'
        ]
        
        # If most of the text is company/header info, skip it
        company_matches = sum(1 for indicator in company_indicators if indicator in combined_text)
        if company_matches > 0 and len(meaningful_fields) <= 2:
            return True
        
        # Check for currency-only rows or incomplete data
        currency_only = all(re.match(r'^[£$€]?[\d.,]+[£$€]?%?$', v) or len(v) <= 3 for v in meaningful_fields)
        if currency_only and len(meaningful_fields) <= 3:
            return True
        
        return False
    
    def _parse_text_for_line_items(self, text: str) -> List[Dict[str, Any]]:
        """Parse text to extract line items when tables are not available."""
        line_items = []
        
        # Pattern to match lines that look like: "Description ... $amount"
        # This is a fallback method for when tables aren't properly detected
        line_pattern = r'(.{10,}?)\s+\$?([0-9,]+\.?\d{0,2})\s*$'
        
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if len(line) < 15:  # Skip short lines
                continue
                
            match = re.search(line_pattern, line)
            if match:
                description = match.group(1).strip()
                amount = match.group(2)
                
                # Filter out lines that look like totals or headers
                skip_patterns = ['total', 'subtotal', 'tax', 'discount', 'grand', 'due', 'balance', 'page']
                if any(pattern in description.lower() for pattern in skip_patterns):
                    continue
                
                if len(description) > 5:  # Reasonable description length
                    line_items.append({
                        'description': description,
                        'amount': amount,
                        'source': 'text_parsing'
                    })
        
        return line_items
    
    def _extract_amount_from_text(self, text: str) -> Optional[str]:
        """Extract a monetary amount from text."""
        if not text:
            return None
        
        # Remove common currency symbols and clean up
        cleaned = re.sub(r'[^\d.,\-]', '', str(text))
        
        # Look for number patterns
        amount_pattern = r'([0-9,]+\.?\d{0,2})'
        match = re.search(amount_pattern, cleaned)
        
        if match:
            return match.group(1)
        return None
    
    def _extract_number_from_text(self, text: str) -> Optional[str]:
        """Extract a number (quantity) from text."""
        if not text:
            return None
        
        # Look for simple numbers
        number_pattern = r'(\d+(?:\.\d+)?)'
        match = re.search(number_pattern, str(text))
        
        if match:
            return match.group(1)
        return None
    
    def _clean_line_item(self, line_item: Dict[str, Any], headers: List[str]) -> Optional[Dict[str, Any]]:
        """Clean and validate a line item, fixing common issues."""
        
        # Skip if no meaningful content
        content_fields = [v for k, v in line_item.items() 
                         if k not in ['page', 'table_number', 'row_number'] and v and str(v).strip()]
        
        if len(content_fields) < 2:  # Need at least description and one other field
            return None
        
        cleaned = {
            'page': line_item.get('page'),
            'table_number': line_item.get('table_number'),
            'row_number': line_item.get('row_number')
        }
        
        # Look for description field and extract quantity if embedded
        description = ""
        quantity = ""
        
        # Find the description field (longest text field or specifically named)
        desc_candidates = []
        for key, value in line_item.items():
            if key in ['page', 'table_number', 'row_number']:
                continue
            if isinstance(value, str) and len(value.strip()) > 5:
                if 'description' in key.lower() or 'quantit' in key.lower():
                    desc_candidates.append((key, value, 100))  # High priority
                else:
                    desc_candidates.append((key, value, len(value)))  # Priority by length
        
        if desc_candidates:
            # Sort by priority (score)
            desc_candidates.sort(key=lambda x: x[2], reverse=True)
            key, desc_text, _ = desc_candidates[0]
            
            # Extract quantity from description if embedded (e.g., "Service Name 14.00")
            import re
            qty_match = re.search(r'\s+(\d+\.?\d*)\s*$', desc_text)
            if qty_match:
                quantity = qty_match.group(1)
                description = desc_text[:qty_match.start()].strip()
            else:
                description = desc_text.strip()
        
        # Look for explicit quantity in other fields
        if not quantity:
            for key, value in line_item.items():
                if key in ['page', 'table_number', 'row_number']:
                    continue
                if 'y' == key.lower() or 'qty' in key.lower() or 'quantity' in key.lower():
                    if isinstance(value, str) and value.strip().replace('.', '').isdigit():
                        quantity = value.strip()
                        break
        
        # Extract other fields
        unit_price = ""
        amount = ""
        vat = ""
        
        for key, value in line_item.items():
            if key in ['page', 'table_number', 'row_number']:
                continue
            
            value_str = str(value).strip() if value else ""
            key_lower = key.lower()
            
            if 'price' in key_lower or 'unit' in key_lower:
                if re.match(r'^\d+\.?\d*$', value_str):
                    unit_price = value_str
            elif 'amount' in key_lower or 'total' in key_lower:
                if re.match(r'^\d+\.?\d*$', value_str):
                    amount = value_str
            elif 'vat' in key_lower or 'tax' in key_lower:
                vat = value_str
        
        # Validate line item - must have description and at least one financial field
        if description and len(description) > 3:
            # Check if this looks like a real line item
            has_financial_data = bool(quantity or unit_price or amount)
            
            # Additional validation - check if description looks like a product/service
            desc_upper = description.upper()
            invalid_patterns = [
                'LIMITED', 'LTD', 'CORP', 'COMPANY', 'INC',
                'TECHNOLOGY', 'SOLUTIONS', 'SERVICES',
                'ADDRESS', 'PHONE', 'EMAIL', 'FAX',
                'ANEXIAN', 'ILICOMM',
                # Single words that aren't products
                r'^\w{1,8}$'  # Single short words
            ]
            
            is_invalid_desc = any(
                pattern in desc_upper if not pattern.startswith('^') 
                else re.match(pattern, desc_upper) 
                for pattern in invalid_patterns
            )
            
            # Only accept if we have financial data and a valid description
            if has_financial_data and not is_invalid_desc:
                cleaned['description'] = description
                if quantity:
                    cleaned['quantity'] = quantity
                if unit_price:
                    cleaned['unit_price'] = unit_price
                if amount:
                    cleaned['amount'] = amount
                if vat:
                    cleaned['vat'] = vat
                
                return cleaned
        
        return None

    def _extract_line_items_from_data(self, tables: List[Dict], text: str) -> List[Dict[str, Any]]:
        """Extract line items from tables and text data."""
        line_items = []
        
        # First try table extraction
        if tables:
            for table_info in tables:
                items = self._extract_from_single_table(table_info)
                line_items.extend(items)
        
        # If no line items found, try text parsing
        if not line_items and text:
            items = self._parse_text_for_line_items(text)
            for item in items:
                item['source'] = 'text_parsing'
                normalized = self._normalize_line_item(item)
                if normalized:
                    line_items.append(normalized)
        
        return line_items
    
    def _extract_from_single_table(self, table_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract line items from a single table."""
        line_items = []
        
        page = table_info['page']
        table_number = table_info['table_number'] 
        rows = table_info['data']
        
        if not rows or len(rows) < 2:
            return line_items
        
        # Find headers - look in first few rows
        headers = None
        header_row_idx = None
        
        for i, row in enumerate(rows[:5]):
            if self._looks_like_headers(row):
                headers = [str(cell).strip().lower() if cell else f'column_{j}' for j, cell in enumerate(row)]
                header_row_idx = i
                break
        
        if not headers:
            return line_items
        
        # Extract data rows after headers
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            
            # Skip empty rows
            if not any(cell and str(cell).strip() for cell in row):
                continue
            
            # Create line item dictionary
            line_item = {
                'page': page,
                'table_number': table_number,
                'row_number': row_idx
            }
            
            # Map row data to headers
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers) and cell_value:
                    header = headers[col_idx]
                    line_item[header] = str(cell_value).strip()
            
            # Basic validation - need at least 2 meaningful fields
            meaningful_fields = [v for k, v in line_item.items() 
                               if k not in ['page', 'table_number', 'row_number'] and v]
            
            if len(meaningful_fields) >= 2:
                # Apply filters
                if not self._is_total_row(line_item):
                    # Normalize the line item
                    normalized = self._normalize_line_item(line_item)
                    if normalized:
                        line_items.append(normalized)
        
        return line_items
    
    def _looks_like_headers(self, row: List[Any]) -> bool:
        """Check if a row looks like table headers."""
        if not row:
            return False
        
        # Convert to strings and check for header-like words
        text_cells = [str(cell).strip().lower() for cell in row if cell and str(cell).strip()]
        
        if len(text_cells) < 2:
            return False
        
        # Look for common invoice table headers
        header_indicators = [
            'description', 'desc', 'item', 'product', 'service',
            'quantity', 'qty', 'quantit', 'amount', 'price', 'rate',
            'unit', 'total', 'vat', 'tax'
        ]
        
        matches = 0
        for cell in text_cells:
            if any(indicator in cell for indicator in header_indicators):
                matches += 1
        
        # Consider it headers if at least 2 cells contain header indicators
        return matches >= 2
    
    def _normalize_line_item(self, line_item: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Normalize a line item into standard fields."""
        
        normalized = {
            'page': line_item.get('page'),
            'table_number': line_item.get('table_number'),
            'row_number': line_item.get('row_number')
        }
        
        # Find description (longest meaningful text field)
        description = ""
        for key, value in line_item.items():
            if key in ['page', 'table_number', 'row_number']:
                continue
            if isinstance(value, str) and len(value.strip()) > 5:
                if 'description' in key or len(value) > len(description):
                    description = value.strip()
        
        if not description or len(description) < 4:
            return None
        
        # ---------------------------------------------
        # NEW: Extract trailing numeric fields embedded
        # ---------------------------------------------
        import re
        quantity = ""
        unit_price = ""
        vat = ""
        
        # Split description into tokens so we can look backwards
        tokens = description.split()
        cleaned_tokens = tokens.copy()
        
        # Walk from end, capturing VAT, unit price, quantity if present
        while cleaned_tokens:
            token = cleaned_tokens[-1]
            if re.match(r'^\d{1,3}%$', token):
                vat = token
                cleaned_tokens.pop()
                continue
            if re.match(r'^\d+\.\d{2}$', token):
                if not unit_price:
                    unit_price = token
                    cleaned_tokens.pop()
                    continue
                elif not quantity:  # second numeric => quantity (can be integer or decimal)
                    quantity = token
                    cleaned_tokens.pop()
                    continue
            if re.match(r'^\d+\.?\d*$', token):
                if not quantity:
                    quantity = token
                    cleaned_tokens.pop()
                    continue
            # If token doesn't match any pattern, break
            break
        
        description_clean = ' '.join(cleaned_tokens).strip()
        
        # Parse quantity from other columns if still missing
        if not quantity:
            for key, value in line_item.items():
                if key in ['page', 'table_number', 'row_number']:
                    continue
                if 'qty' in key.lower() or 'quantit' in key.lower() or key.lower() == 'y':
                    if isinstance(value, str) and value.strip().replace('.', '').isdigit():
                        quantity = value.strip()
                        break
        
        # Parse unit_price, amount, vat from other headers if present
        amount = ""
        for key, value in line_item.items():
            if key in ['page', 'table_number', 'row_number']:
                continue
            value_str = str(value).strip()
            key_lower = key.lower()
            if not unit_price and ('price' in key_lower or 'unit' in key_lower):
                if re.match(r'^\d+\.?\d*$', value_str):
                    unit_price = value_str
            elif not amount and ('amount' in key_lower or 'total' in key_lower):
                if re.match(r'^\d+\.?\d*$', value_str):
                    amount = value_str
            elif not vat and ('vat' in key_lower or 'tax' in key_lower):
                vat = value_str
        
        # If amount still missing but numeric in line_item['amount'] maybe captured earlier
        if not amount and 'amount' in line_item:
            amount = line_item['amount']
        
        # Validate that we have a proper line item (needs description and at least amount or unit_price)
        if description_clean and (unit_price or amount):
            normalized['description'] = description_clean
            if quantity:
                normalized['quantity'] = quantity
            if unit_price:
                normalized['unit_price'] = unit_price
            if amount:
                normalized['amount'] = amount
            if vat:
                normalized['vat'] = vat
            return normalized
        
        return None


class ExcelExporter:
    """Handles exporting extracted data to Excel format."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_summary_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create a summary sheet with overview of all processed files."""
        ws = workbook.create_sheet("Summary", 0)
        
        # Headers
        headers = ['Filename', 'Pages', 'Tables Found', 'Line Items Found', 'Has Text', 'File Size (KB)', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row, data in enumerate(all_data, 2):
            filepath = Path(data['filepath'])
            file_size = filepath.stat().st_size / 1024 if filepath.exists() else 0
            
            values = [
                data['filename'],
                data['metadata'].get('pages', 0),
                len(data['tables']),
                len(data.get('line_items', [])),
                'Yes' if data['text'].strip() else 'No',
                f"{file_size:.1f}",
                'Error' if 'error' in data else 'Success'
            ]
            
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def create_line_items_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create a sheet with extracted line items from all PDFs, preserving original column structure."""
        ws = workbook.create_sheet("Line Items")
        
        # Collect all unique column names from all line items
        all_column_names = set()
        # Exclude these internal/meta fields from the Excel headers
        standard_columns = ['filename', 'page', 'table_number', 'row_number', 'source']
        
        for data in all_data:
            if 'error' not in data and 'line_items' in data and data['line_items']:
                for line_item in data['line_items']:
                    for key in line_item.keys():
                        if key not in standard_columns:
                            all_column_names.add(key)
        
        # Desired preferred order after Filename
        preferred_order = ['description', 'quantity', 'unit_price', 'vat', 'amount']
        
        # Create comprehensive header list in preferred order
        headers = ['Filename']
        original_columns = list(all_column_names)
        
        # Add preferred columns if present
        for col in preferred_order:
            if col in original_columns:
                headers.append(col.replace('_', ' ').title())
                original_columns.remove(col)
        
        # Append any remaining columns alphabetically
        for col_name in sorted(original_columns):
            headers.append(col_name.replace('_', ' ').title())
        
        # Finally add Source column
        headers.append('Source')
        
        # Create header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        current_row = 2
        
        # Fill in the data
        for data in all_data:
            if 'error' not in data and 'line_items' in data and data['line_items']:
                for line_item in data['line_items']:
                    values = [data['filename']]
                    
                    # Add values in preferred order first
                    for col in preferred_order:
                        if col in line_item:
                            values.append(line_item.get(col, ''))
                    # Then remaining columns (same order as headers after preferred)
                    for col_name in sorted(original_columns):
                        values.append(line_item.get(col_name, ''))
                    
                    # Add source
                    values.append(line_item.get('source', 'table_parsing'))
                    
                    # Write to Excel
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        # Enable text wrapping for longer content
                        if isinstance(value, str) and len(value) > 30:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    current_row += 1
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(ws.columns, 1):
            if column[0].value:  # If column has a header
                max_length = 0
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # Set column width with reasonable limits
                if max_length > 40:
                    ws.column_dimensions[column[0].column_letter].width = 40
                elif max_length < 10:
                    ws.column_dimensions[column[0].column_letter].width = 12
                else:
                    ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # Add statistics at the top
        if current_row > 2:
            total_items = current_row - 2
            total_columns = len(original_columns)
            ws.insert_rows(1)
            stats_cell = ws.cell(row=1, column=1, 
                               value=f"Total Line Items: {total_items} | Unique Columns Found: {total_columns}")
            stats_cell.font = Font(bold=True, size=12)
            stats_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # Merge across first few columns (filename + first data cols)
            merge_end = min(3, len(headers))
            if merge_end > 1:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end)
    
    def create_text_data_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create a sheet with extracted text and parsed invoice data."""
        ws = workbook.create_sheet("Text Data")
        
        # Headers for structured data
        headers = ['Filename', 'Invoice Number', 'Date', 'Vendor', 'Total Amount', 'Full Text']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        extractor = PDFDataExtractor()
        
        for row, data in enumerate(all_data, 2):
            if 'error' not in data and data['text']:
                invoice_data = extractor.extract_invoice_data(data['text'])
                
                values = [
                    data['filename'],
                    invoice_data.get('invoice_number', ''),
                    invoice_data.get('date', ''),
                    invoice_data.get('vendor', ''),
                    invoice_data.get('total_amount', ''),
                    data['text'][:1000] + '...' if len(data['text']) > 1000 else data['text']
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col == 6:  # Full text column
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def create_tables_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create a sheet with all extracted tables."""
        ws = workbook.create_sheet("Tables")
        
        current_row = 1
        
        for data in all_data:
            if 'error' not in data and data['tables']:
                for table_info in data['tables']:
                    # Table header
                    header_text = f"File: {data['filename']} | Page: {table_info['page']} | Table: {table_info['table_number']}"
                    cell = ws.cell(row=current_row, column=1, value=header_text)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    current_row += 1
                    
                    # Table data
                    table_data = table_info['data']
                    for table_row in table_data:
                        for col, cell_value in enumerate(table_row, 1):
                            if cell_value is not None:
                                ws.cell(row=current_row, column=col, value=str(cell_value))
                        current_row += 1
                    
                    current_row += 1  # Empty row between tables
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def export_to_excel(self, all_data: List[Dict[str, Any]], output_file: Path) -> None:
        """
        Export all extracted data to Excel file with multiple sheets.
        
        Args:
            all_data: List of extracted data dictionaries
            output_file: Path for the output Excel file
        """
        try:
            workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Create sheets
            self.create_summary_sheet(workbook, all_data)
            self.create_line_items_sheet(workbook, all_data)
            self.create_text_data_sheet(workbook, all_data)
            self.create_tables_sheet(workbook, all_data)
            
            # Save workbook
            workbook.save(output_file)
            self.logger.info(f"Excel file saved: {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
            raise


def setup_logging(verbose: bool = False) -> None:
    """Setup logging configuration."""
    # Configure root logger
    root_logger = logging.getLogger()
    
    # Clear any existing handlers
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)
    
    # Set our application's logging level
    app_level = logging.DEBUG if verbose else logging.INFO
    
    # Create formatters
    console_formatter = logging.Formatter('%(levelname)s - %(message)s')
    file_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    # Console handler (less verbose)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(app_level)
    console_handler.setFormatter(console_formatter)
    
    # File handler (more detailed)
    file_handler = logging.FileHandler('pdf_converter.log')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(file_formatter)
    
    # Configure our application logger
    app_logger = logging.getLogger(__name__)
    app_logger.setLevel(app_level)
    app_logger.addHandler(console_handler)
    app_logger.addHandler(file_handler)
    
    # Suppress noisy third-party loggers unless in verbose mode
    if not verbose:
        # Suppress PDF processing library debug messages
        logging.getLogger('pdfminer').setLevel(logging.WARNING)
        logging.getLogger('pdfplumber').setLevel(logging.WARNING)
        logging.getLogger('PIL').setLevel(logging.WARNING)
        logging.getLogger('pypdfium2').setLevel(logging.WARNING)
        logging.getLogger('fontTools').setLevel(logging.WARNING)
    else:
        # In verbose mode, still suppress the most noisy ones
        logging.getLogger('pdfminer.psparser').setLevel(logging.WARNING)
        logging.getLogger('pdfminer.pdfinterp').setLevel(logging.WARNING)


def find_pdf_files(input_dir: Path) -> List[Path]:
    """Find all PDF files in the input directory."""
    pdf_files = set()  # Use set to avoid duplicates
    for pattern in ['*.pdf', '*.PDF']:
        pdf_files.update(input_dir.glob(pattern))
    
    # Convert back to sorted list
    unique_files = sorted(list(pdf_files))
    return unique_files


def main():
    """Main function to run the PDF to Excel converter."""
    parser = argparse.ArgumentParser(
        description="Extract data from multiple PDF files and export to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --input-dir ./pdfs --output-file results.xlsx
  %(prog)s --input-dir ./invoices --extract-tables --output-file invoice_data.xlsx --verbose
        """
    )
    
    parser.add_argument(
        '--input-dir', '-i',
        type=Path,
        default=Path('./pdfs'),
        help='Directory containing PDF files (default: ./pdfs)'
    )
    
    parser.add_argument(
        '--output-file', '-o',
        type=Path,
        default=Path('./extracted_data.xlsx'),
        help='Output Excel file path (default: ./extracted_data.xlsx)'
    )
    
    parser.add_argument(
        '--extract-tables',
        action='store_true',
        default=True,
        help='Extract tables from PDFs (default: True)'
    )
    
    parser.add_argument(
        '--extract-text',
        action='store_true',
        default=True,
        help='Extract text from PDFs (default: True)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging'
    )
    
    args = parser.parse_args()
    
    # Setup logging
    setup_logging(args.verbose)
    logger = logging.getLogger(__name__)
    
    # Validate input directory
    if not args.input_dir.exists():
        logger.error(f"Input directory does not exist: {args.input_dir}")
        sys.exit(1)
    
    # Find PDF files
    pdf_files = find_pdf_files(args.input_dir)
    if not pdf_files:
        logger.error(f"No PDF files found in: {args.input_dir}")
        sys.exit(1)
    
    logger.info(f"Found {len(pdf_files)} PDF files to process")
    
    # Initialize extractor and exporter
    extractor = PDFDataExtractor(
        extract_tables=args.extract_tables,
        extract_text=args.extract_text
    )
    exporter = ExcelExporter()
    
    # Process all PDF files
    all_extracted_data = []
    for i, pdf_file in enumerate(pdf_files, 1):
        print(f"Processing {i}/{len(pdf_files)}: {pdf_file.name}")
        logger.info(f"Processing {i}/{len(pdf_files)}: {pdf_file.name}")
        extracted_data = extractor.extract_from_pdf(pdf_file)
        all_extracted_data.append(extracted_data)
        
        # Show quick progress indicator
        if 'line_items' in extracted_data:
            line_item_count = len(extracted_data['line_items'])
            if line_item_count > 0:
                print(f"  ✓ Found {line_item_count} line items")
            else:
                print(f"  - No line items detected")
        else:
            print(f"  - Processing completed")
    
    # Export to Excel
    print(f"\nExporting data to Excel: {args.output_file}")
    logger.info(f"Exporting data to Excel: {args.output_file}")
    exporter.export_to_excel(all_extracted_data, args.output_file)
    
    # Summary
    successful_extractions = len([d for d in all_extracted_data if 'error' not in d])
    total_line_items = sum(len(d.get('line_items', [])) for d in all_extracted_data)
    
    print(f"\n🎉 Processing Complete!")
    print(f"📊 Files processed: {successful_extractions}/{len(pdf_files)}")
    print(f"📋 Total line items found: {total_line_items}")
    print(f"💾 Results saved to: {args.output_file}")
    
    logger.info(f"Processing complete! {successful_extractions}/{len(pdf_files)} files processed successfully")
    logger.info(f"Total line items extracted: {total_line_items}")
    logger.info(f"Results saved to: {args.output_file}")


if __name__ == "__main__":
    main() 