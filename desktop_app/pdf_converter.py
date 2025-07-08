#!/usr/bin/env python3
"""
Simplified PDF to Excel Converter for Streamlit App
===================================================

Essential PDF processing functionality for the desktop application.
"""

import os
import sys
import logging
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

# Third-party imports
try:
    import pandas as pd
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please install required packages: pip install pandas pdfplumber openpyxl")
    sys.exit(1)


class PDFDataExtractor:
    """Handles extraction of data from PDF files."""
    
    def __init__(self, extract_tables: bool = True, extract_text: bool = True):
        self.extract_tables = extract_tables
        self.extract_text = extract_text
        self.logger = logging.getLogger(__name__)
    
    def extract_from_pdf(self, pdf_path: Path) -> Dict[str, Any]:
        """
        Extract data from a single PDF file.
        
        Args:
            pdf_path: Path to the PDF file
            
        Returns:
            Dictionary containing extracted data
        """
        extracted_data = {
            'filename': pdf_path.name,
            'filepath': str(pdf_path),
            'text': '',
            'tables': [],
            'metadata': {},
            'extraction_date': datetime.now().isoformat()
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Extract metadata
                if pdf.metadata:
                    extracted_data['metadata'] = {
                        'title': pdf.metadata.get('Title', ''),
                        'author': pdf.metadata.get('Author', ''),
                        'creator': pdf.metadata.get('Creator', ''),
                        'creation_date': str(pdf.metadata.get('CreationDate', '')),
                        'pages': len(pdf.pages)
                    }
                
                all_text = []
                all_tables = []
                
                for page_num, page in enumerate(pdf.pages, 1):
                    if self.extract_text:
                        page_text = page.extract_text()
                        if page_text:
                            all_text.append(f"Page {page_num}:\n{page_text}\n")
                    
                    if self.extract_tables:
                        # Extract tables
                        tables = []
                        
                        # First try default table extraction
                        default_tables = page.extract_tables()
                        if default_tables:
                            for i, table in enumerate(default_tables):
                                if table:  # Skip empty tables
                                    tables.append(table)
                                    all_tables.append({
                                        'page': page_num,
                                        'table_number': i + 1,
                                        'data': table,
                                        'rows': len(table),
                                        'columns': len(table[0]) if table else 0
                                    })
                        
                        # If no tables found with default method, try text-based detection
                        if not tables:
                            table_settings = {
                                "vertical_strategy": "text",
                                "horizontal_strategy": "text", 
                                "intersection_tolerance": 5,
                                "min_words_vertical": 1,
                                "min_words_horizontal": 1
                            }
                            text_tables = page.extract_tables(table_settings)
                            
                            for i, table in enumerate(text_tables):
                                if table and len(table) > 1:  # Skip empty or single-row tables
                                    # Filter out mostly empty tables
                                    meaningful_rows = 0
                                    for row in table:
                                        if row and any(cell and str(cell).strip() for cell in row):
                                            meaningful_rows += 1
                                    
                                    if meaningful_rows >= 2:  # At least 2 meaningful rows
                                        tables.append(table)
                                        all_tables.append({
                                            'page': page_num,
                                            'table_number': i + 1,
                                            'data': table,
                                            'rows': len(table),
                                            'columns': len(table[0]) if table else 0,
                                            'detection_method': 'text-based'
                                        })
                
                extracted_data['text'] = '\n'.join(all_text)
                extracted_data['tables'] = all_tables
                
                # Extract line items from tables and text
                extracted_data['line_items'] = self._extract_line_items_from_data(all_tables, extracted_data['text'])
                
                # Debug info
                self.logger.info(f"Extracted {len(all_tables)} tables and {len(extracted_data['line_items'])} line items from {pdf_path.name}")
                
        except Exception as e:
            self.logger.error(f"Error extracting data from {pdf_path}: {str(e)}")
            extracted_data['error'] = str(e)
        
        return extracted_data
    
    def _extract_line_items_from_data(self, tables: List[Dict], text: str) -> List[Dict[str, Any]]:
        """Extract line items from tables and text data."""
        line_items = []
        
        # First try table extraction
        if tables:
            for table_info in tables:
                items = self._extract_from_single_table(table_info)
                line_items.extend(items)
        
        # If no line items found, try text parsing
        if not line_items and text:
            items = self._parse_text_for_line_items(text)
            for item in items:
                item['source'] = 'text_parsing'
                normalized = self._normalize_line_item(item)
                if normalized:
                    line_items.append(normalized)
        
        return line_items
    
    def _extract_from_single_table(self, table_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract line items from a single table."""
        line_items = []
        
        page = table_info['page']
        table_number = table_info['table_number'] 
        rows = table_info['data']
        
        if not rows or len(rows) < 2:
            return line_items
        
        # Find headers - look in first few rows
        headers = None
        header_row_idx = None
        
        for i, row in enumerate(rows[:5]):
            if self._looks_like_headers(row):
                headers = [str(cell).strip().lower() if cell else f'column_{j}' for j, cell in enumerate(row)]
                header_row_idx = i
                break
        
        if not headers:
            return line_items
        
        # Extract data rows after headers
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            
            # Skip empty rows
            if not any(cell and str(cell).strip() for cell in row):
                continue
            
            # Create line item dictionary
            line_item = {
                'page': page,
                'table_number': table_number,
                'row_number': row_idx
            }
            
            # Map row data to headers
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers) and cell_value:
                    header = headers[col_idx]
                    line_item[header] = str(cell_value).strip()
            
            # Basic validation - need at least 2 meaningful fields
            meaningful_fields = [v for k, v in line_item.items() 
                               if k not in ['page', 'table_number', 'row_number'] and v]
            
            if len(meaningful_fields) >= 2:
                # Apply filters
                if not self._is_total_row(line_item):
                    # Normalize the line item
                    normalized = self._normalize_line_item(line_item)
                    if normalized:
                        line_items.append(normalized)
        
        return line_items
    
    def _looks_like_headers(self, row: List[Any]) -> bool:
        """Check if a row looks like table headers."""
        if not row:
            return False
        
        # Convert to strings and check for header-like words
        text_cells = [str(cell).strip().lower() for cell in row if cell and str(cell).strip()]
        
        if len(text_cells) < 2:
            return False
        
        # Look for common invoice table headers
        header_indicators = [
            'description', 'desc', 'item', 'product', 'service',
            'quantity', 'qty', 'quantit', 'amount', 'price', 'rate',
            'unit', 'total', 'vat', 'tax'
        ]
        
        matches = 0
        for cell in text_cells:
            if any(indicator in cell for indicator in header_indicators):
                matches += 1
        
        # Consider it headers if at least 2 cells contain header indicators
        return matches >= 2
    
    def _normalize_line_item(self, line_item: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Normalize a line item into standard fields."""
        
        normalized = {
            'page': line_item.get('page'),
            'table_number': line_item.get('table_number'),
            'row_number': line_item.get('row_number')
        }
        
        # Find description (longest meaningful text field)
        description = ""
        for key, value in line_item.items():
            if key in ['page', 'table_number', 'row_number']:
                continue
            if isinstance(value, str) and len(value.strip()) > 5:
                if 'description' in key or len(value) > len(description):
                    description = value.strip()
        
        if not description or len(description) < 4:
            return None
        
        # Extract trailing numeric fields embedded
        import re
        quantity = ""
        unit_price = ""
        vat = ""
        
        # Split description into tokens so we can look backwards
        tokens = description.split()
        cleaned_tokens = tokens.copy()
        
        # Walk from end, capturing VAT, unit price, quantity if present
        while cleaned_tokens:
            token = cleaned_tokens[-1]
            if re.match(r'^\d{1,3}%$', token):
                vat = token
                cleaned_tokens.pop()
                continue
            if re.match(r'^\d+\.\d{2}$', token):
                if not unit_price:
                    unit_price = token
                    cleaned_tokens.pop()
                    continue
                elif not quantity:  # second numeric => quantity (can be integer or decimal)
                    quantity = token
                    cleaned_tokens.pop()
                    continue
            if re.match(r'^\d+\.?\d*$', token):
                if not quantity:
                    quantity = token
                    cleaned_tokens.pop()
                    continue
            # If token doesn't match any pattern, break
            break
        
        description_clean = ' '.join(cleaned_tokens).strip()
        
        # Parse quantity from other columns if still missing
        if not quantity:
            for key, value in line_item.items():
                if key in ['page', 'table_number', 'row_number']:
                    continue
                if 'qty' in key.lower() or 'quantit' in key.lower() or key.lower() == 'y':
                    if isinstance(value, str) and value.strip().replace('.', '').isdigit():
                        quantity = value.strip()
                        break
        
        # Parse unit_price, amount, vat from other headers if present
        amount = ""
        for key, value in line_item.items():
            if key in ['page', 'table_number', 'row_number']:
                continue
            value_str = str(value).strip()
            key_lower = key.lower()
            if not unit_price and ('price' in key_lower or 'unit' in key_lower):
                if re.match(r'^\d+\.?\d*$', value_str):
                    unit_price = value_str
            elif not amount and ('amount' in key_lower or 'total' in key_lower):
                if re.match(r'^\d+\.?\d*$', value_str):
                    amount = value_str
            elif not vat and ('vat' in key_lower or 'tax' in key_lower):
                vat = value_str
        
        # If amount still missing but numeric in line_item['amount'] maybe captured earlier
        if not amount and 'amount' in line_item:
            amount = line_item['amount']
        
        # Validate that we have a proper line item (needs description and at least amount or unit_price)
        if description_clean and (unit_price or amount):
            normalized['description'] = description_clean
            if quantity:
                normalized['quantity'] = quantity
            if unit_price:
                normalized['unit_price'] = unit_price
            if amount:
                normalized['amount'] = amount
            if vat:
                normalized['vat'] = vat
            return normalized
        
        return None
    
    def _parse_text_for_line_items(self, text: str) -> List[Dict[str, Any]]:
        """Parse text to extract line items when tables are not available."""
        line_items = []
        
        # Pattern to match lines that look like: "Description ... $amount"
        # This is a fallback method for when tables aren't properly detected
        line_pattern = r'(.{10,}?)\s+\$?([0-9,]+\.?\d{0,2})\s*$'
        
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if len(line) < 15:  # Skip short lines
                continue
                
            match = re.search(line_pattern, line)
            if match:
                description = match.group(1).strip()
                amount = match.group(2)
                
                # Filter out lines that look like totals or headers
                skip_patterns = ['total', 'subtotal', 'tax', 'discount', 'grand', 'due', 'balance', 'page']
                if any(pattern in description.lower() for pattern in skip_patterns):
                    continue
                
                if len(description) > 5:  # Reasonable description length
                    line_items.append({
                        'description': description,
                        'amount': amount,
                        'source': 'text_parsing'
                    })
        
        return line_items
    
    def _extract_number(self, text: str) -> Optional[float]:
        """Extract a number from text."""
        if not text:
            return None
        
        # Remove common non-numeric characters but keep decimal points
        cleaned = re.sub(r'[^\d.,]', '', str(text))
        
        # Handle different decimal formats
        if ',' in cleaned and '.' in cleaned:
            # Assume comma is thousands separator
            cleaned = cleaned.replace(',', '')
        elif ',' in cleaned and cleaned.count(',') == 1:
            # Assume comma is decimal separator
            cleaned = cleaned.replace(',', '.')
        
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None
    
    def _extract_amount(self, text: str) -> Optional[float]:
        """Extract a monetary amount from text."""
        if not text:
            return None
        
        # Remove currency symbols and extract number
        cleaned = re.sub(r'[£$€¥₹]', '', str(text))
        return self._extract_number(cleaned)
    
    def _is_total_row(self, row_data: Dict[str, Any]) -> bool:
        """Check if a row represents a total/summary rather than a line item."""
        
        # Convert all values to strings for checking
        text_values = []
        for key, value in row_data.items():
            if key not in ['page', 'table_number', 'row_number'] and value:
                text_values.append(str(value).strip().upper())
        
        combined_text = ' '.join(text_values)
        
        # Check for total/summary indicators
        total_indicators = [
            'TOTAL', 'SUBTOTAL', 'SUB-TOTAL', 'GRAND TOTAL',
            'VAT', 'TAX', 'NET', 'GROSS', 'BALANCE',
            'AMOUNT DUE', 'INVOICE TOTAL', 'FINAL TOTAL',
            'SHIPPING', 'DELIVERY', 'DISCOUNT',
            'TOTA', 'OTAL',  # Partial matches for split text
        ]
        
        # Check if any value contains total indicators
        for indicator in total_indicators:
            if indicator in combined_text:
                return True
        
        # Check for rows with mostly empty fields or single words
        meaningful_fields = [v for v in text_values if len(v) > 2]
        if len(meaningful_fields) <= 1:
            return True
        
        # Check for company/header information
        company_indicators = [
            'LIMITED', 'LTD', 'LLC', 'CORP', 'CORPORATION', 'INC',
            'TECHNOLOGY', 'SOLUTIONS', 'SERVICES', 'GROUP',
            'COMPANY', 'ANEXIAN'
        ]
        
        # If most of the text is company/header info, skip it
        company_matches = sum(1 for indicator in company_indicators if indicator in combined_text)
        if company_matches > 0 and len(meaningful_fields) <= 2:
            return True
        
        # Check for currency-only rows or incomplete data
        currency_only = all(re.match(r'^[£$€]?[\d.,]+[£$€]?%?$', v) or len(v) <= 3 for v in meaningful_fields)
        if currency_only and len(meaningful_fields) <= 3:
            return True
        
        return False
    
    def extract_invoice_data(self, text: str) -> Dict[str, str]:
        """
        Extract common invoice fields from text using regex patterns.
        
        Args:
            text: Extracted text from PDF
            
        Returns:
            Dictionary with extracted invoice fields
        """
        invoice_data = {}
        
        # Common patterns for invoice data
        patterns = {
            'invoice_number': [
                r'invoice\s*#?\s*:?\s*([A-Z0-9\-_]+)',
                r'inv\s*#?\s*:?\s*([A-Z0-9\-_]+)',
                r'invoice\s*number\s*:?\s*([A-Z0-9\-_]+)'
            ],
            'date': [
                r'date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'invoice\s*date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'
            ],
            'total_amount': [
                r'total\s*:?\s*\$?([0-9,]+\.?\d*)',
                r'amount\s*due\s*:?\s*\$?([0-9,]+\.?\d*)',
                r'grand\s*total\s*:?\s*\$?([0-9,]+\.?\d*)'
            ],
            'vendor': [
                r'from\s*:?\s*([^\n]+)',
                r'vendor\s*:?\s*([^\n]+)',
                r'company\s*:?\s*([^\n]+)'
            ]
        }
        
        for field, pattern_list in patterns.items():
            for pattern in pattern_list:
                match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if match:
                    invoice_data[field] = match.group(1).strip()
                    break
        
        return invoice_data


class ExcelExporter:
    """Handles export of extracted data to Excel files."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def export_to_excel(self, all_data: List[Dict[str, Any]], output_file: Path) -> None:
        """Export extracted data to Excel file with multiple sheets."""
        try:
            workbook = Workbook()
            
            # Remove default sheet
            if workbook.worksheets:
                workbook.remove(workbook.active)
            
            # Create sheets
            self.create_summary_sheet(workbook, all_data)
            self.create_line_items_sheet(workbook, all_data)
            self.create_text_data_sheet(workbook, all_data)
            self.create_tables_sheet(workbook, all_data)
            
            # Save workbook
            workbook.save(output_file)
            self.logger.info(f"Excel file saved successfully: {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {e}")
            raise
    
    def create_summary_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create summary sheet with file processing overview."""
        ws = workbook.create_sheet("Summary")
        
        # Headers
        headers = ['Filename', 'Status', 'Pages', 'Tables Found', 'Line Items', 'File Size']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row, data in enumerate(all_data, 2):
            status = "Error" if 'error' in data else "Success"
            file_path = Path(data.get('filepath', ''))
            file_size = file_path.stat().st_size if file_path.exists() else 0
            
            values = [
                data['filename'],
                status,
                data.get('metadata', {}).get('pages', 0),
                len(data.get('tables', [])),
                len(data.get('line_items', [])),
                f"{file_size / 1024:.1f} KB"
            ]
            
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def create_line_items_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create line items sheet with extracted invoice items."""
        ws = workbook.create_sheet("Line Items")
        
        # Headers
        headers = ['Filename', 'Description', 'Quantity', 'Unit Price', 'Amount', 'Source']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Data rows
        row = 2
        total_items = 0
        
        for data in all_data:
            if 'error' not in data:
                line_items = data.get('line_items', [])
                total_items += len(line_items)
                
                if line_items:
                    for item in line_items:
                        values = [
                            data['filename'],
                            item.get('description', ''),
                            item.get('quantity', ''),
                            item.get('unit_price', ''),
                            item.get('amount', ''),
                            item.get('source', '')
                        ]
                        
                        for col, value in enumerate(values, 1):
                            ws.cell(row=row, column=col, value=value)
                        
                        row += 1
                else:
                    # Add a row indicating no line items found for this file
                    values = [
                        data['filename'],
                        'No line items detected in this file',
                        '',
                        '',
                        '',
                        'No tables with structured data found'
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        if col == 2:  # Description column
                            cell.font = Font(italic=True)
                    
                    row += 1
        
        # Add summary if no items found at all
        if total_items == 0:
            ws.cell(row=row, column=1, value="No line items detected in any files")
            ws.cell(row=row, column=2, value="The PDFs may not contain structured table data or may be image-based")
            for col in range(1, 3):
                ws.cell(row=row, column=col).font = Font(bold=True, italic=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def create_text_data_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create a sheet with extracted text and parsed invoice data."""
        ws = workbook.create_sheet("Text Data")
        
        # Headers for structured data
        headers = ['Filename', 'Invoice Number', 'Date', 'Vendor', 'Total Amount', 'Full Text']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        extractor = PDFDataExtractor()
        
        for row, data in enumerate(all_data, 2):
            if 'error' not in data and data['text']:
                invoice_data = extractor.extract_invoice_data(data['text'])
                
                values = [
                    data['filename'],
                    invoice_data.get('invoice_number', ''),
                    invoice_data.get('date', ''),
                    invoice_data.get('vendor', ''),
                    invoice_data.get('total_amount', ''),
                    data['text'][:1000] + '...' if len(data['text']) > 1000 else data['text']
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col == 6:  # Full text column
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def create_tables_sheet(self, workbook: Workbook, all_data: List[Dict[str, Any]]) -> None:
        """Create tables sheet with detected table structures."""
        ws = workbook.create_sheet("Tables")
        
        # Headers
        headers = ['Filename', 'Page', 'Table #', 'Rows', 'Columns', 'Table Data (Preview)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Data rows
        row = 2
        for data in all_data:
            if 'error' not in data:
                for table in data.get('tables', []):
                    # Create a preview of the table data
                    table_data = table.get('data', [])
                    preview = []
                    for table_row in table_data[:3]:  # First 3 rows
                        if table_row:
                            row_text = ' | '.join(str(cell)[:20] if cell else '' for cell in table_row[:5])  # First 5 columns
                            preview.append(row_text)
                    
                    values = [
                        data['filename'],
                        table.get('page', ''),
                        table.get('table_number', ''),
                        table.get('rows', ''),
                        table.get('columns', ''),
                        '\n'.join(preview)
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        if col == 6:  # Table data column
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60) 